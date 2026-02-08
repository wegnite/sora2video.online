import argparse
import json
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Optional

import requests
from openpyxl import load_workbook


KW_HEADERS = [
    "keyword",
    "keywords",
    "palavra-chave",  # pt-BR
    "schlüsselwort",  # de
    "キーワード",       # ja
]


def translate_google_public(text: str, target: str = "zh-CN") -> str:
    base = "https://translate.googleapis.com/translate_a/single"
    params = {
        "client": "gtx",
        "sl": "auto",
        "tl": target,
        "dt": "t",
        "q": text,
    }
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        r = requests.get(base, params=params, headers=headers, timeout=20)
        r.raise_for_status()
        data = r.json()
        return data[0][0][0]
    except Exception:
        return ""


def find_keyword_column(header_cells: List[Optional[str]]) -> int:
    lowered = [str(h).strip().strip('"').lower() if h is not None else "" for h in header_cells]
    for cand in KW_HEADERS:
        if cand in lowered:
            return lowered.index(cand) + 1  # 1-based index for openpyxl
    # fallback: if there is a column literally named 'Keyword' with case differences
    for i, h in enumerate(header_cells, start=1):
        if h and str(h).strip().lower().startswith("keyword"):
            return i
    raise ValueError(f"Could not find Keyword column in header: {header_cells}")


def load_cache(cache_path: Path) -> Dict[str, str]:
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cache(cache_path: Path, cache: Dict[str, str]) -> None:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = cache_path.with_suffix(cache_path.suffix + ".tmp")
    tmp.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
    tmp.replace(cache_path)


def translate_unique(terms: List[str], concurrency: int, cache: Dict[str, str]) -> Dict[str, str]:
    uniq = []
    for t in terms:
        key = (t or "").strip()
        if key and key not in cache and key not in uniq:
            uniq.append(key)
    if not uniq:
        return cache

    def worker(t: str) -> (str, str):
        # small backoff spacing
        txt = translate_google_public(t, target="zh-CN")
        if not txt:
            time.sleep(0.2)
            txt = translate_google_public(t, target="zh-CN")
        return t, txt

    with ThreadPoolExecutor(max_workers=max(1, concurrency)) as ex:
        futures = {ex.submit(worker, t): t for t in uniq}
        for fut in as_completed(futures):
            t, zh = fut.result()
            cache[t] = zh
    return cache


def process_file(path: Path, out_suffix: str, concurrency: int, cache: Dict[str, str], cache_path: Path) -> Path:
    wb = load_workbook(path)
    ws = wb.active
    # Read header row (1-based)
    header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    try:
        kw_col = find_keyword_column(header)
    except ValueError:
        # Assume second column is keyword if header missing
        kw_col = 2 if ws.max_column >= 2 else 1

    # Ensure third column is Chinese meaning; if header already exists, reuse col index
    zh_header = "中文释义"
    # Insert at 3 if not already present
    existing_headers = [str(h).strip() if h is not None else "" for h in header]
    if len(existing_headers) < 3 or existing_headers[2] != zh_header:
        ws.insert_cols(3)
        ws.cell(row=1, column=3).value = zh_header

    # Collect terms
    terms = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=kw_col).value
        terms.append(str(val).strip() if val is not None else "")

    # Translate with cache + concurrency
    translate_unique(terms, concurrency, cache)
    # Persist cache frequently
    save_cache(cache_path, cache)

    # Fill worksheet
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=kw_col).value
        key = str(val).strip() if val is not None else ""
        ws.cell(row=row, column=3).value = cache.get(key, "")

    out_path = path.with_name(path.stem + out_suffix + path.suffix)
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser(description="Translate Excel keyword files and insert Chinese meanings as column 3")
    ap.add_argument("input", help="Path to an .xlsx file or a directory containing .xlsx files")
    ap.add_argument("--pattern", default="*.xlsx", help="Glob pattern when input is a directory")
    ap.add_argument("--suffix", default="_with_cn", help="Suffix for output files before extension")
    ap.add_argument("-n", "--concurrency", type=int, default=6, help="Concurrent translation workers")
    ap.add_argument("--cache", default="data/translate_cache.json", help="Path to translation cache JSON")
    args = ap.parse_args()

    input_path = Path(args.input)
    cache_path = Path(args.cache)
    cache = load_cache(cache_path)

    targets: List[Path] = []
    if input_path.is_dir():
        targets = sorted(input_path.glob(args.pattern))
    else:
        targets = [input_path]

    if not targets:
        print("No .xlsx files found.")
        return

    for p in targets:
        if p.name.startswith("~$"):  # skip temporary Excel lock files
            continue
        print(f"Translating: {p}")
        out = process_file(p, args.suffix, args.concurrency, cache, cache_path)
        print(f" -> Wrote: {out}")
        # Save cache after each file
        save_cache(cache_path, cache)


if __name__ == "__main__":
    main()

