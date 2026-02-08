import re
import sys
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


URL = "https://r.jina.ai/http://www.toolify.ai/de/Best-AI-Tools-revenue"


def fetch_text(url: str) -> str:
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    return resp.text


def extract_description_lines(markdown_text: str) -> list[str]:
    lines = []
    for raw in markdown_text.splitlines():
        line = raw.strip()
        if not line:
            continue
        # skip the header meta lines if present
        if line.startswith("Title:") or line.startswith("URL Source:") or line.startswith("Markdown Content:"):
            continue
        # Heuristic: description lines contain a revenue figure like 5.9B / 327.1M etc
        if not re.search(r"\b\d+(?:\.\d+)?[MB]\b", line):
            continue
        # Take substring after the LAST revenue token on the line
        m_iter = list(re.finditer(r"\b\d+(?:\.\d+)?[MB]\b", line))
        if not m_iter:
            continue
        last = m_iter[-1]
        desc = line[last.end():].strip()
        if desc:
            lines.append(desc)
    return lines


def tokenize(text: str) -> list[str]:
    # Replace common separators and hyphens with space, keep unicode letters
    # Keep case as-is; drop standalone numbers
    cleaned = (
        text.replace("—", " ")
            .replace("–", " ")
            .replace("-", " ")
            .replace("/", " ")
            .replace("|", " ")
    )
    # Remove punctuation except word characters and spaces (keep unicode letters)
    cleaned = re.sub(r"[\.,!?:;\(\)\[\]\{\}\"'`]+", " ", cleaned)
    # Collapse whitespace
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    tokens = []
    for tok in cleaned.split(" "):
        if not tok:
            continue
        # Skip pure numbers
        if re.fullmatch(r"\d+", tok):
            continue
        # Skip stray separators
        if tok in {"|", "•", "·", "•"}:
            continue
        tokens.append(tok)
    return tokens


def main(out_path: Path) -> None:
    html = fetch_text(URL)
    # Jina reader returns a markdown-like block under "Markdown Content:"; extract everything after that label
    # Fallback: just use full text
    text = html
    # Attempt to isolate markdown content block
    mk_idx = text.find("Markdown Content:")
    if mk_idx != -1:
        text = text[mk_idx + len("Markdown Content:"):]

    descriptions = extract_description_lines(text)
    tokens: list[str] = []
    for desc in descriptions:
        tokens.extend(tokenize(desc))

    # Write to Excel (one column, include duplicates in original order)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "keywords"
    ws.cell(row=1, column=1, value="keyword")
    for i, tok in enumerate(tokens, start=2):
        ws.cell(row=i, column=1, value=tok)
    wb.save(out_path)


if __name__ == "__main__":
    out = Path("data/toolify_revenue_keywords.xlsx")
    if len(sys.argv) > 1:
        out = Path(sys.argv[1])
    main(out)
